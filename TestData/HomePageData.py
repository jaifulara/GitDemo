import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname" :"J", "lastname" :"F", "Contact" :"1234", "Email" :"Gmail.com", "Message" :"Hey"},
                            {"firstname" :"Jay", "lastname" :"Ful", "Contact" :"3211", "Email" :"mail.com", "Message" :"How"},
                            {"firstname" :"H", "lastname" :"U", "Contact" :"3322", "Email" :"g@mail.com", "Message" :"Are"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("C:\\Users\\Jaideep Fulara\\OneDrive\\Documents\\PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]
